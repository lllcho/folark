"""XLSX 文件处理实现。"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

from app.plugins.builtin_plugin.utils import render_template
from app.plugins.core import PreviewResult

if TYPE_CHECKING:
    from app.plugins.core import PluginContext

logger = logging.getLogger(__name__)


def extract(ctx: PluginContext) -> str | None:
    """从 XLSX 文件提取文本。"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(ctx.file_path, read_only=True, data_only=True)
        text_parts = []

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        text_parts.append(str(cell.value))

        wb.close()

        text = "\n".join(text_parts).strip()
        return text if text else None
    except Exception as e:
        logger.warning("XLSX 文本提取失败 %s: %s", ctx.file_path, e)
        return None


def preview(ctx: PluginContext) -> PreviewResult:
    """使用模板渲染 XLSX 预览页面。"""
    html = render_template(
        "xlsx_preview.html",
        {
            "file_name": ctx.file_name,
            "title": ctx.title or ctx.file_name,
            "file_url": ctx.file_url,
        },
    )
    return PreviewResult.from_html(html)
